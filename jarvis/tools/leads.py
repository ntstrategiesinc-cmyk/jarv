"""Lead-management tools backed by an Excel workbook (leads.xlsx).

Reads (list/find/summarize) are safe and auto-run. Adding a lead is a write, so it's marked
consequential and routes through the confirmation gate. The store is plain and human-readable:
open leads.xlsx in Excel any time to inspect or correct it.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ..config import Config
from .base import Tool, ToolResult

COLUMNS = ["name", "email", "phone", "business", "source", "status", "notes", "created_at"]


class LeadsStore:
    def __init__(self, path: Path, sheet_name: str):
        self.path = Path(path)
        self.sheet = sheet_name

    def _open_or_create(self) -> Workbook:
        if self.path.exists():
            wb = load_workbook(self.path)
            if self.sheet not in wb.sheetnames:
                ws = wb.create_sheet(self.sheet)
                ws.append(COLUMNS)
            return wb
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet
        ws.append(COLUMNS)
        return wb

    def all(self) -> list[dict]:
        """Return every lead as a dict. Empty list if the workbook doesn't exist yet
        (reads never create the file)."""
        if not self.path.exists():
            return []
        wb = load_workbook(self.path, read_only=True)
        if self.sheet not in wb.sheetnames:
            return []
        ws = wb[self.sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h) if h is not None else "" for h in rows[0]]
        leads: list[dict] = []
        for row in rows[1:]:
            if row is None or all(c is None for c in row):
                continue
            record = {header[i]: ("" if i >= len(row) or row[i] is None else row[i]) for i in range(len(header))}
            leads.append(record)
        return leads

    def add(self, fields: dict) -> dict:
        """Append one lead and save. Raises on I/O problems (e.g. file open in Excel);
        the tool handler turns that into a plain-language error."""
        wb = self._open_or_create()
        ws = wb[self.sheet]
        record = {c: str(fields.get(c, "") or "") for c in COLUMNS}
        if not record["status"]:
            record["status"] = "new"
        record["created_at"] = datetime.now().isoformat(timespec="seconds")
        ws.append([record[c] for c in COLUMNS])
        wb.save(self.path)
        return record


def _fmt_lead(lead: dict) -> str:
    bits = [lead.get("name", "")]
    for k in ("business", "status", "email", "phone"):
        v = lead.get(k, "")
        if v:
            bits.append(f"{k}={v}")
    return " | ".join(b for b in bits if b)


def build_leads_tools(config: Config) -> list[Tool]:
    store = LeadsStore(config.leads_workbook_path, config.leads_sheet_name)

    def leads_add(args: dict) -> ToolResult:
        name = (args.get("name") or "").strip()
        if not name:
            return ToolResult.error("A lead needs at least a name.")
        try:
            record = store.add(args)
        except PermissionError:
            return ToolResult.error(
                "Couldn't write to the spreadsheet — it may be open in Excel. "
                "Close it and try again. Nothing was saved."
            )
        except Exception as e:  # surface to the model, don't crash
            return ToolResult.error(f"Couldn't save the lead: {e}")
        return ToolResult.success(f"Added lead: {_fmt_lead(record)} (created {record['created_at']}).")

    def leads_list(args: dict) -> ToolResult:
        try:
            leads = store.all()
        except Exception as e:
            return ToolResult.error(f"Couldn't read the spreadsheet: {e}")
        status = (args.get("status") or "").strip().lower()
        business = (args.get("business") or "").strip().lower()
        if status:
            leads = [l for l in leads if str(l.get("status", "")).lower() == status]
        if business:
            leads = [l for l in leads if business in str(l.get("business", "")).lower()]
        if not leads:
            return ToolResult.success("No matching leads.")
        lines = [f"{i + 1}. {_fmt_lead(l)}" for i, l in enumerate(leads)]
        return ToolResult.success(f"{len(leads)} lead(s):\n" + "\n".join(lines))

    def leads_find(args: dict) -> ToolResult:
        query = (args.get("query") or "").strip().lower()
        if not query:
            return ToolResult.error("Provide a search query.")
        try:
            leads = store.all()
        except Exception as e:
            return ToolResult.error(f"Couldn't read the spreadsheet: {e}")
        hits = [
            l for l in leads
            if any(query in str(l.get(k, "")).lower() for k in ("name", "email", "phone", "business", "notes"))
        ]
        if not hits:
            return ToolResult.success(f"No leads match '{query}'.")
        lines = [f"{i + 1}. {_fmt_lead(l)}" for i, l in enumerate(hits)]
        return ToolResult.success(f"{len(hits)} match(es):\n" + "\n".join(lines))

    def leads_summarize(args: dict) -> ToolResult:
        try:
            leads = store.all()
        except Exception as e:
            return ToolResult.error(f"Couldn't read the spreadsheet: {e}")
        if not leads:
            return ToolResult.success("No leads recorded yet.")
        by_status: dict[str, int] = {}
        by_business: dict[str, int] = {}
        for l in leads:
            by_status[str(l.get("status") or "unknown")] = by_status.get(str(l.get("status") or "unknown"), 0) + 1
            biz = str(l.get("business") or "unspecified")
            by_business[biz] = by_business.get(biz, 0) + 1
        status_str = ", ".join(f"{k}: {v}" for k, v in sorted(by_status.items()))
        biz_str = ", ".join(f"{k}: {v}" for k, v in sorted(by_business.items()))
        return ToolResult.success(
            f"{len(leads)} total lead(s). By status — {status_str}. By business — {biz_str}."
        )

    return [
        Tool(
            name="leads_add",
            description=(
                "Add a new sales lead to the spreadsheet. Use when the user mentions a new "
                "prospect or contact to track. A write action."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "The lead's name (required)."},
                    "email": {"type": "string"},
                    "phone": {"type": "string"},
                    "business": {"type": "string", "description": "Which of the owner's businesses this lead is for."},
                    "source": {"type": "string", "description": "Where the lead came from (web, referral, event, ...)."},
                    "status": {"type": "string", "description": "Stage: new, contacted, qualified, won, lost. Defaults to new."},
                    "notes": {"type": "string"},
                },
                "required": ["name"],
            },
            handler=leads_add,
            needs_confirmation=True,
        ),
        Tool(
            name="leads_list",
            description="List leads, optionally filtered by status or business. Read-only.",
            input_schema={
                "type": "object",
                "properties": {
                    "status": {"type": "string", "description": "Only leads with this status."},
                    "business": {"type": "string", "description": "Only leads for this business."},
                },
            },
            handler=leads_list,
            needs_confirmation=False,
        ),
        Tool(
            name="leads_find",
            description="Search leads by text across name, email, phone, business, and notes. Read-only.",
            input_schema={
                "type": "object",
                "properties": {"query": {"type": "string", "description": "Text to search for."}},
                "required": ["query"],
            },
            handler=leads_find,
            needs_confirmation=False,
        ),
        Tool(
            name="leads_summarize",
            description="Summarize the lead pipeline: totals and counts by status and business. Read-only.",
            input_schema={"type": "object", "properties": {}},
            handler=leads_summarize,
            needs_confirmation=False,
        ),
    ]
